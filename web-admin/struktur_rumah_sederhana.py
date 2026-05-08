"""
Tabulasi Perhitungan Struktur Rumah Sederhana 2 Lantai
DENGAN FORMULA EXCEL - HITUNG OTOMATIS
Based on SNI 2847:2019, SNI 1726:2019, SNI 1727:2020, SNI 8460:2017
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def create_calculated_structure():
    wb = openpyxl.Workbook()

    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color='FFFFFF')
    input_font = Font(bold=True, color='0000FF')
    label_font = Font(bold=True, size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    thick_border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    calc_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    result_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    def sh(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    def si(cell):
        cell.font = input_font
        cell.fill = input_fill
        cell.alignment = center
        cell.border = thick_border

    def sl(cell):
        cell.font = label_font
        cell.alignment = left
        cell.border = thin_border

    def sc(cell):
        cell.fill = calc_fill
        cell.border = thin_border
        cell.alignment = center

    def sr(cell):
        cell.fill = result_fill
        cell.border = thin_border
        cell.alignment = center

    def st(cell):
        cell.fill = ok_fill
        cell.border = thin_border
        cell.alignment = center

    # ============================================================
    # SHEET 1: INPUT - DATA BANGUNAN & MATERIAL
    # ============================================================
    ws1 = wb.active
    ws1.title = "01_Input"

    ws1.merge_cells('A1:F1')
    ws1['A1'] = "INPUT DATA BANGUNAN & MATERIAL"
    ws1['A1'].font = Font(bold=True, size=16, color='4472C4')
    ws1['A1'].alignment = center

    ws1.merge_cells('A2:F2')
    ws1['A2'] = "Ubah nilai di sel kuning untuk menghitung otomatis"
    ws1['A2'].font = Font(italic=True, color='666666')
    ws1['A2'].alignment = center

    # Data Bangunan
    ws1.merge_cells('A4:D4')
    ws1['A4'] = "DATA BANGUNAN"
    ws1['A4'].font = title_font
    ws1['A4'].alignment = center

    sh(ws1.cell(row=5, column=1, value="Parameter"))
    sh(ws1.cell(row=5, column=2, value="Nilai"))
    sh(ws1.cell(row=5, column=3, value="Satuan"))
    sh(ws1.cell(row=5, column=4, value="Keterangan"))

    input_data = [
        ('Panjang (Px)', 8, 'm', 'Arah X'),
        ('Lebar (Py)', 6, 'm', 'Arah Y'),
        ('Tinggi Lantai 1', 3.5, 'm', 'Tinggi bersih'),
        ('Tinggi Lantai 2', 3.0, 'm', 'Tinggi bersih'),
        ('Jumlah Lantai', 2, '', 'Bangunan bertingkat'),
        ('Tebal Plat Lantai', 12, 'cm', 'Sesuai SNI 2847'),
        ('Tebal Plat Atap', 10, 'cm', 'Beton ringan'),
        ('Tebal Dinding', 10, 'cm', 'Batako'),
    ]

    for i, (param, val, satuan, ket) in enumerate(input_data, 6):
        sl(ws1.cell(row=i, column=1, value=param))
        cell = ws1.cell(row=i, column=2, value=val)
        si(cell)
        cell.number_format = '0.00'
        sc(ws1.cell(row=i, column=3, value=satuan))
        sc(ws1.cell(row=i, column=4, value=ket))

    # Beban
    ws1.merge_cells('A15:D15')
    ws1['A15'] = "BEBAN DESAIN (SNI 1727:2020)"
    ws1['A15'].font = title_font
    ws1['A15'].alignment = center

    sh(ws1.cell(row=16, column=1, value="Jenis Beban"))
    sh(ws1.cell(row=16, column=2, value="Nilai"))
    sh(ws1.cell(row=16, column=3, value="Satuan"))
    sh(ws1.cell(row=16, column=4, value="Referensi"))

    beban_data = [
        ('Beban Hidup Lantai', 1.92, 'kN/m2', 'SNI 1727:2020'),
        ('Beban Hidup Atap', 0.96, 'kN/m2', 'SNI 1727:2020'),
        ('Beban Mati Plat Lantai', 3.0, 'kN/m2', 'Berat sendiri'),
        ('Beban Mati Plat Atap', 2.5, 'kN/m2', 'Berat sendiri'),
        ('Berat Dinding Batako', 1.8, 'kN/m2', 'γ=18 kN/m3'),
    ]

    for i, (jenis, val, satuan, ref) in enumerate(beban_data, 17):
        sl(ws1.cell(row=i, column=1, value=jenis))
        cell = ws1.cell(row=i, column=2, value=val)
        si(cell)
        cell.number_format = '0.00'
        sc(ws1.cell(row=i, column=3, value=satuan))
        sc(ws1.cell(row=i, column=4, value=ref))

    # Material
    ws1.merge_cells('A23:D23')
    ws1['A23'] = "PROPERTI MATERIAL (SNI 2847:2019)"
    ws1['A23'].font = title_font
    ws1['A23'].alignment = center

    sh(ws1.cell(row=24, column=1, value="Material"))
    sh(ws1.cell(row=24, column=2, value="Nilai"))
    sh(ws1.cell(row=24, column=3, value="Satuan"))
    sh(ws1.cell(row=24, column=4, value="Pasal"))

    mat_data = [
        ('Kuat Tekan Beton (fc)', 19.3, 'MPa', 'Pasal 6.1'),
        ('Mutu Baja Tulangan (fy)', 420, 'MPa', 'Pasal 6.5'),
        ('Berat Beton (γc)', 24, 'kN/m3', 'Pasal 6.10'),
        ('Faktor Reduksi Lentur (φ)', 0.65, '-', 'Pasal 21.2.1'),
        ('Faktor Reduksi Geser (φV)', 0.75, '-', 'Pasal 21.2.1'),
        ('Selimut Beton (cc)', 25, 'mm', 'Tabel 20.6.1.3.1'),
    ]

    for i, (mat, val, satuan, pasal) in enumerate(mat_data, 25):
        sl(ws1.cell(row=i, column=1, value=mat))
        cell = ws1.cell(row=i, column=2, value=val)
        si(cell)
        cell.number_format = '0.00'
        sc(ws1.cell(row=i, column=3, value=satuan))
        sc(ws1.cell(row=i, column=4, value=pasal))

    ws1.column_dimensions['A'].width = 26
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 20

    # ============================================================
    # SHEET 2: SLOOF - PERHITUNGAN OTOMATIS
    # ============================================================
    ws2 = wb.create_sheet("02_Sloof")

    ws2.merge_cells('A1:I1')
    ws2['A1'] = "PERHITUNGAN SLOOF - OTOMATIS"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = center

    ws2.merge_cells('A2:I2')
    ws2['A2'] = "Sesuai SNI 2847:2019 Pasal 9"
    ws2['A2'].alignment = center

    # Headers
    headers = ['No', 'Lokasi', 'Bentang (m)', 'q Dinding\n(kN/m)', 'q Total\n(kN/m)', 'Mu\n(kNm)', 'Vu\n(kN)', 'As Perlu\n(mm2)', 'Status']
    for col, h in enumerate(headers, 1):
        sh(ws2.cell(row=4, column=col, value=h))
        ws2.row_dimensions[4].height = 30

    # Sloof 1: Melintang
    ws2.cell(row=5, column=1, value=1)
    sc(ws2.cell(row=5, column=1))
    ws2.cell(row=5, column=2, value='Sloof Melintang')
    sc(ws2.cell(row=5, column=2))

    ws2.cell(row=5, column=3, value='=01_Input!B6')
    sc(ws2.cell(row=5, column=3))
    ws2.cell(row=5, column=3).number_format = '0.00'

    # q dinding = tebal(m) * tinggi Lt.1 * 18 kN/m3
    ws2.cell(row=5, column=4, value='=(15/100)*01_Input!B8*18')
    sc(ws2.cell(row=5, column=4))
    ws2.cell(row=5, column=4).number_format = '0.00'

    # q total
    ws2.cell(row=5, column=5, value='=D5+0.5')
    sc(ws2.cell(row=5, column=5))
    ws2.cell(row=5, column=5).number_format = '0.00'

    # Mu = 1/8 * q * L^2
    ws2.cell(row=5, column=6, value='=1/8*E5*C5^2')
    sr(ws2.cell(row=5, column=6))
    ws2.cell(row=5, column=6).number_format = '0.00'

    # Vu = 0.5 * q * L
    ws2.cell(row=5, column=7, value='=0.5*E5*C5')
    sr(ws2.cell(row=5, column=7))
    ws2.cell(row=5, column=7).number_format = '0.00'

    # As perlu
    ws2.cell(row=5, column=8, value='=F5*1000000/(01_Input!B28*01_Input!B29*(30-2.5-0.4)*10)')
    sr(ws2.cell(row=5, column=8))
    ws2.cell(row=5, column=8).number_format = '0.00'

    ws2.cell(row=5, column=9, value='=IF(H5<398,"OK","NOK")')
    st(ws2.cell(row=5, column=9))

    # Sloof 2: Memanjang
    ws2.cell(row=6, column=1, value=2)
    sc(ws2.cell(row=6, column=1))
    ws2.cell(row=6, column=2, value='Sloof Memanjang')
    sc(ws2.cell(row=6, column=2))

    ws2.cell(row=6, column=3, value='=01_Input!B7')
    sc(ws2.cell(row=6, column=3))
    ws2.cell(row=6, column=3).number_format = '0.00'

    ws2.cell(row=6, column=4, value='=(15/100)*01_Input!B8*18')
    sc(ws2.cell(row=6, column=4))
    ws2.cell(row=6, column=4).number_format = '0.00'

    ws2.cell(row=6, column=5, value='=D6+0.5')
    sc(ws2.cell(row=6, column=5))
    ws2.cell(row=6, column=5).number_format = '0.00'

    ws2.cell(row=6, column=6, value='=1/8*E6*C6^2')
    sr(ws2.cell(row=6, column=6))
    ws2.cell(row=6, column=6).number_format = '0.00'

    ws2.cell(row=6, column=7, value='=0.5*E6*C6')
    sr(ws2.cell(row=6, column=7))
    ws2.cell(row=6, column=7).number_format = '0.00'

    ws2.cell(row=6, column=8, value='=F6*1000000/(01_Input!B28*01_Input!B29*(30-2.5-0.4)*10)')
    sr(ws2.cell(row=6, column=8))
    ws2.cell(row=6, column=8).number_format = '0.00'

    ws2.cell(row=6, column=9, value='=IF(H6<398,"OK","NOK")')
    st(ws2.cell(row=6, column=9))

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 12
    ws2.column_dimensions['H'].width = 12
    ws2.column_dimensions['I'].width = 10

    # ============================================================
    # SHEET 3: KOLOM - PERHITUNGAN OTOMATIS
    # ============================================================
    ws3 = wb.create_sheet("03_Kolom")

    ws3.merge_cells('A1:J1')
    ws3['A1'] = "PERHITUNGAN KOLOM - OTOMATIS"
    ws3['A1'].font = title_font
    ws3['A1'].alignment = center

    ws3.merge_cells('A2:J2')
    ws3['A2'] = "SNI 2847:2019 Pasal 10 & 21.6 - SRPMM"
    ws3['A2'].alignment = center

    kol_headers = ['No', 'Tipe', 'Atribut\n(m2)', 'Pu (kN)', 'Ag perlu\n(mm2)', 'Ag ada\n(mm2)', 'ρ (%)', 'Tul. Utama', 'Sengkang', 'Status']
    for col, h in enumerate(kol_headers, 1):
        sh(ws3.cell(row=4, column=col, value=h))
        ws3.row_dimensions[4].height = 30

    # K1 - pojok (25% area)
    ws3.cell(row=5, column=1, value=1)
    sc(ws3.cell(row=5, column=1))
    ws3.cell(row=5, column=2, value='K1 (Pojok)')
    sc(ws3.cell(row=5, column=2))

    ws3.cell(row=5, column=3, value='=01_Input!B6*01_Input!B7/4')
    sc(ws3.cell(row=5, column=3))
    ws3.cell(row=5, column=3).number_format = '0.00'

    # Pu = At*(beban plat + beban dinding) + berat kolom
    ws3.cell(row=5, column=4, value='=C5*((01_Input!B11/100)*24 + (01_Input!B13/100)*01_Input!B8*18)/1000 + 0.3*0.3*01_Input!B8*24/1000')
    sr(ws3.cell(row=5, column=4))
    ws3.cell(row=5, column=4).number_format = '0.00'

    # Ag perlu
    ws3.cell(row=5, column=5, value='=D5*1000/(01_Input!B28*0.8*01_Input!B25*1000)')
    sr(ws3.cell(row=5, column=5))
    ws3.cell(row=5, column=5).number_format = '0.00'

    ws3.cell(row=5, column=6, value=90000)
    sc(ws3.cell(row=5, column=6))

    ws3.cell(row=5, column=7, value='=(4*132.7)/(30*30)*100')
    sr(ws3.cell(row=5, column=7))
    ws3.cell(row=5, column=7).number_format = '0.00'

    ws3.cell(row=5, column=8, value='4D13')
    sc(ws3.cell(row=5, column=8))
    ws3.cell(row=5, column=9, value='D10-100')
    sc(ws3.cell(row=5, column=9))
    ws3.cell(row=5, column=10, value='=IF(F5>E5,"OK","NOK")')
    st(ws3.cell(row=5, column=10))

    # K2 - tepi (50% area)
    ws3.cell(row=6, column=1, value=2)
    sc(ws3.cell(row=6, column=1))
    ws3.cell(row=6, column=2, value='K2 (Tepi)')
    sc(ws3.cell(row=6, column=2))

    ws3.cell(row=6, column=3, value='=01_Input!B6*01_Input!B7/2')
    sc(ws3.cell(row=6, column=3))
    ws3.cell(row=6, column=3).number_format = '0.00'

    ws3.cell(row=6, column=4, value='=C6*((01_Input!B11/100)*24 + (01_Input!B13/100)*01_Input!B8*18)/1000 + 0.3*0.3*01_Input!B8*24/1000')
    sr(ws3.cell(row=6, column=4))
    ws3.cell(row=6, column=4).number_format = '0.00'

    ws3.cell(row=6, column=5, value='=D6*1000/(01_Input!B28*0.8*01_Input!B25*1000)')
    sr(ws3.cell(row=6, column=5))
    ws3.cell(row=6, column=5).number_format = '0.00'

    ws3.cell(row=6, column=6, value=90000)
    sc(ws3.cell(row=6, column=6))

    ws3.cell(row=6, column=7, value='=(4*132.7)/(30*30)*100')
    sr(ws3.cell(row=6, column=7))
    ws3.cell(row=6, column=7).number_format = '0.00'

    ws3.cell(row=6, column=8, value='4D13')
    sc(ws3.cell(row=6, column=8))
    ws3.cell(row=6, column=9, value='D10-100')
    sc(ws3.cell(row=6, column=9))
    ws3.cell(row=6, column=10, value='=IF(F6>E6,"OK","NOK")')
    st(ws3.cell(row=6, column=10))

    # K3 - tengah (100% area)
    ws3.cell(row=7, column=1, value=3)
    sc(ws3.cell(row=7, column=1))
    ws3.cell(row=7, column=2, value='K3 (Tengah)')
    sc(ws3.cell(row=7, column=2))

    ws3.cell(row=7, column=3, value='=01_Input!B6*01_Input!B7')
    sc(ws3.cell(row=7, column=3))
    ws3.cell(row=7, column=3).number_format = '0.00'

    ws3.cell(row=7, column=4, value='=C7*((01_Input!B11/100)*24 + (01_Input!B13/100)*01_Input!B8*18)/1000 + 0.3*0.3*01_Input!B8*24/1000')
    sr(ws3.cell(row=7, column=4))
    ws3.cell(row=7, column=4).number_format = '0.00'

    ws3.cell(row=7, column=5, value='=D7*1000/(01_Input!B28*0.8*01_Input!B25*1000)')
    sr(ws3.cell(row=7, column=5))
    ws3.cell(row=7, column=5).number_format = '0.00'

    ws3.cell(row=7, column=6, value=90000)
    sc(ws3.cell(row=7, column=6))

    ws3.cell(row=7, column=7, value='=(4*132.7)/(30*30)*100')
    sr(ws3.cell(row=7, column=7))
    ws3.cell(row=7, column=7).number_format = '0.00'

    ws3.cell(row=7, column=8, value='4D16')
    sc(ws3.cell(row=7, column=8))
    ws3.cell(row=7, column=9, value='D10-100')
    sc(ws3.cell(row=7, column=9))
    ws3.cell(row=7, column=10, value='=IF(F7>E7,"OK","NOK")')
    st(ws3.cell(row=7, column=10))

    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 12
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 10
    ws3.column_dimensions['H'].width = 10
    ws3.column_dimensions['I'].width = 10
    ws3.column_dimensions['J'].width = 10

    # ============================================================
    # SHEET 4: BALOK - PERHITUNGAN OTOMATIS
    # ============================================================
    ws4 = wb.create_sheet("04_Balok")

    ws4.merge_cells('A1:L1')
    ws4['A1'] = "PERHITUNGAN BALOK - OTOMATIS"
    ws4['A1'].font = title_font
    ws4['A1'].alignment = center

    ws4.merge_cells('A2:L2')
    ws4['A2'] = "SNI 2847:2019 Pasal 8 & 21.5 - SRPMM"
    ws4['A2'].alignment = center

    balok_headers = ['No', 'Tipe', 'Bentang\n(m)', 'qd\n(kN/m)', 'ql\n(kN/m)', 'Mu\n(kNm)', 'Vu\n(kN)', 'd eff\n(mm)', 'As perlu\n(mm2)', 'As ada\n(mm2)', 'ρ (%)', 'Status']
    for col, h in enumerate(balok_headers, 1):
        sh(ws4.cell(row=4, column=col, value=h))
        ws4.row_dimensions[4].height = 30

    # B1 - bentang Px
    ws4.cell(row=5, column=1, value=1)
    sc(ws4.cell(row=5, column=1))
    ws4.cell(row=5, column=2, value='B1 (Lantai)')
    sc(ws4.cell(row=5, column=2))
    ws4.cell(row=5, column=3, value='=01_Input!B6')
    sc(ws4.cell(row=5, column=3))
    ws4.cell(row=5, column=3).number_format = '0.00'

    # qd = berat balok + beban plat + beban dinding
    ws4.cell(row=5, column=4, value='=(25*40*24)/1000000+01_Input!B19/1000*C5+01_Input!B21/2*C5')
    sc(ws4.cell(row=5, column=4))
    ws4.cell(row=5, column=4).number_format = '0.00'

    ws4.cell(row=5, column=5, value='=C5*01_Input!B17/1000')
    sc(ws4.cell(row=5, column=5))
    ws4.cell(row=5, column=5).number_format = '0.00'

    ws4.cell(row=5, column=6, value='=1/11*(D5+E5)*C5^2')
    sr(ws4.cell(row=5, column=6))
    ws4.cell(row=5, column=6).number_format = '0.00'

    ws4.cell(row=5, column=7, value='=0.5*(D5+E5)*C5')
    sr(ws4.cell(row=5, column=7))
    ws4.cell(row=5, column=7).number_format = '0.00'

    ws4.cell(row=5, column=8, value=370)
    sc(ws4.cell(row=5, column=8))

    ws4.cell(row=5, column=9, value='=F5*1000000/(01_Input!B28*01_Input!B29*H5)')
    sr(ws4.cell(row=5, column=9))
    ws4.cell(row=5, column=9).number_format = '0.00'

    ws4.cell(row=5, column=10, value=603)
    sc(ws4.cell(row=5, column=10))

    ws4.cell(row=5, column=11, value='=J5/(25*40)*100')
    sr(ws4.cell(row=5, column=11))
    ws4.cell(row=5, column=11).number_format = '0.00'

    ws4.cell(row=5, column=12, value='=IF(J5>I5,"OK","NOK")')
    st(ws4.cell(row=5, column=12))

    # B2 - bentang Py
    ws4.cell(row=6, column=1, value=2)
    sc(ws4.cell(row=6, column=1))
    ws4.cell(row=6, column=2, value='B2 (Lantai)')
    sc(ws4.cell(row=6, column=2))
    ws4.cell(row=6, column=3, value='=01_Input!B7')
    sc(ws4.cell(row=6, column=3))
    ws4.cell(row=6, column=3).number_format = '0.00'

    ws4.cell(row=6, column=4, value='=(25*40*24)/1000000+01_Input!B19/1000*C6+01_Input!B21/2*C6')
    sc(ws4.cell(row=6, column=4))
    ws4.cell(row=6, column=4).number_format = '0.00'

    ws4.cell(row=6, column=5, value='=C6*01_Input!B17/1000')
    sc(ws4.cell(row=6, column=5))
    ws4.cell(row=6, column=5).number_format = '0.00'

    ws4.cell(row=6, column=6, value='=1/11*(D6+E6)*C6^2')
    sr(ws4.cell(row=6, column=6))
    ws4.cell(row=6, column=6).number_format = '0.00'

    ws4.cell(row=6, column=7, value='=0.5*(D6+E6)*C6')
    sr(ws4.cell(row=6, column=7))
    ws4.cell(row=6, column=7).number_format = '0.00'

    ws4.cell(row=6, column=8, value=370)
    sc(ws4.cell(row=6, column=8))

    ws4.cell(row=6, column=9, value='=F6*1000000/(01_Input!B28*01_Input!B29*H6)')
    sr(ws4.cell(row=6, column=9))
    ws4.cell(row=6, column=9).number_format = '0.00'

    ws4.cell(row=6, column=10, value=402)
    sc(ws4.cell(row=6, column=10))

    ws4.cell(row=6, column=11, value='=J6/(25*40)*100')
    sr(ws4.cell(row=6, column=11))
    ws4.cell(row=6, column=11).number_format = '0.00'

    ws4.cell(row=6, column=12, value='=IF(J6>I6,"OK","NOK")')
    st(ws4.cell(row=6, column=12))

    # B3 - atap
    ws4.cell(row=7, column=1, value=3)
    sc(ws4.cell(row=7, column=1))
    ws4.cell(row=7, column=2, value='B3 (Atap)')
    sc(ws4.cell(row=7, column=2))
    ws4.cell(row=7, column=3, value='=01_Input!B6')
    sc(ws4.cell(row=7, column=3))
    ws4.cell(row=7, column=3).number_format = '0.00'

    ws4.cell(row=7, column=4, value='=(25*40*24)/1000000+C7*(01_Input!B20+01_Input!B21)/1000')
    sc(ws4.cell(row=7, column=4))
    ws4.cell(row=7, column=4).number_format = '0.00'

    ws4.cell(row=7, column=5, value='=C7*01_Input!B18/1000')
    sc(ws4.cell(row=7, column=5))
    ws4.cell(row=7, column=5).number_format = '0.00'

    ws4.cell(row=7, column=6, value='=1/11*(D7+E7)*C7^2')
    sr(ws4.cell(row=7, column=6))
    ws4.cell(row=7, column=6).number_format = '0.00'

    ws4.cell(row=7, column=7, value='=0.5*(D7+E7)*C7')
    sr(ws4.cell(row=7, column=7))
    ws4.cell(row=7, column=7).number_format = '0.00'

    ws4.cell(row=7, column=8, value=370)
    sc(ws4.cell(row=7, column=8))

    ws4.cell(row=7, column=9, value='=F7*1000000/(01_Input!B28*01_Input!B29*H7)')
    sr(ws4.cell(row=7, column=9))
    ws4.cell(row=7, column=9).number_format = '0.00'

    ws4.cell(row=7, column=10, value=402)
    sc(ws4.cell(row=7, column=10))

    ws4.cell(row=7, column=11, value='=J7/(25*40)*100')
    sr(ws4.cell(row=7, column=11))
    ws4.cell(row=7, column=11).number_format = '0.00'

    ws4.cell(row=7, column=12, value='=IF(J7>I7,"OK","NOK")')
    st(ws4.cell(row=7, column=12))

    ws4.column_dimensions['A'].width = 6
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 10
    ws4.column_dimensions['D'].width = 12
    ws4.column_dimensions['E'].width = 10
    ws4.column_dimensions['F'].width = 12
    ws4.column_dimensions['G'].width = 10
    ws4.column_dimensions['H'].width = 10
    ws4.column_dimensions['I'].width = 12
    ws4.column_dimensions['J'].width = 12
    ws4.column_dimensions['K'].width = 10
    ws4.column_dimensions['L'].width = 10

    # ============================================================
    # SHEET 5: PLAT LANTAI - PERHITUNGAN OTOMATIS
    # ============================================================
    ws5 = wb.create_sheet("05_Plat_Lantai")

    ws5.merge_cells('A1:J1')
    ws5['A1'] = "PERHITUNGAN PLAT LANTAI - OTOMATIS"
    ws5['A1'].font = title_font
    ws5['A1'].alignment = center

    plat_headers = ['Panel', 'Lx (m)', 'Ly (m)', 'ly/lx', 'Tipe', 'Mu lap\n(kNm)', 'd eff\n(mm)', 'As perlu\n(mm2/m)', 'As ada\n(mm2/m)', 'Status']
    for col, h in enumerate(plat_headers, 1):
        sh(ws5.cell(row=4, column=col, value=h))
        ws5.row_dimensions[4].height = 30

    # Panel A1
    ws5.cell(row=5, column=1, value='A1')
    sc(ws5.cell(row=5, column=1))
    ws5.cell(row=5, column=2, value='=01_Input!B7')
    sc(ws5.cell(row=5, column=2))
    ws5.cell(row=5, column=2).number_format = '0.00'
    ws5.cell(row=5, column=3, value='=01_Input!B6')
    sc(ws5.cell(row=5, column=3))
    ws5.cell(row=5, column=3).number_format = '0.00'
    ws5.cell(row=5, column=4, value='=C5/B5')
    sc(ws5.cell(row=5, column=4))
    ws5.cell(row=5, column=4).number_format = '0.00'
    ws5.cell(row=5, column=5, value='=IF(D5>2,"Satu Arah","Dua Arah")')
    sc(ws5.cell(row=5, column=5))
    ws5.cell(row=5, column=6, value='=0.058*(01_Input!B19+01_Input!B17)*B5^2')
    sr(ws5.cell(row=5, column=6))
    ws5.cell(row=5, column=6).number_format = '0.00'
    ws5.cell(row=5, column=7, value=100)
    sc(ws5.cell(row=5, column=7))
    ws5.cell(row=5, column=8, value='=F5*1000000/(01_Input!B28*01_Input!B29*G5)')
    sr(ws5.cell(row=5, column=8))
    ws5.cell(row=5, column=8).number_format = '0.00'
    ws5.cell(row=5, column=9, value=523)
    sc(ws5.cell(row=5, column=9))
    ws5.cell(row=5, column=10, value='=IF(I5>H5,"OK","NOK")')
    st(ws5.cell(row=5, column=10))

    ws5.column_dimensions['A'].width = 10
    ws5.column_dimensions['B'].width = 10
    ws5.column_dimensions['C'].width = 10
    ws5.column_dimensions['D'].width = 10
    ws5.column_dimensions['E'].width = 12
    ws5.column_dimensions['F'].width = 12
    ws5.column_dimensions['G'].width = 10
    ws5.column_dimensions['H'].width = 12
    ws5.column_dimensions['I'].width = 12
    ws5.column_dimensions['J'].width = 10

    # ============================================================
    # SHEET 6: FONDASI FOOTPLATE - PERHITUNGAN OTOMATIS
    # ============================================================
    ws6 = wb.create_sheet("06_Fondasi")

    ws6.merge_cells('A1:K1')
    ws6['A1'] = "PERHITUNGAN FONDASI FOOTPLATE - OTOMATIS"
    ws6['A1'].font = title_font
    ws6['A1'].alignment = center

    ws6.merge_cells('A2:K2')
    ws6['A2'] = "SNI 8460:2017 & SNI 2847:2019 Pasal 8.4"
    ws6['A2'].alignment = center

    # Kapasitas tanah
    ws6.cell(row=4, column=1, value="KAPASITAS TANAH (σ)")
    sh(ws6.cell(row=4, column=1))
    ws6.cell(row=4, column=2, value=150)
    si(ws6.cell(row=4, column=2))
    ws6.cell(row=4, column=3, value="kN/m2")
    sc(ws6.cell(row=4, column=3))

    fp_headers = ['Kolom', 'Pu (kN)', 'σ tanah\n(kN/m2)', 'B (m)', 'Luas perlu\n(m2)', 'Luas ada\n(m2)', 'Tekanan\n(kN/m2)', 'Status', 'b0 (mm)', 'φVn (N)', 'Status Pons']
    for col, h in enumerate(fp_headers, 1):
        sh(ws6.cell(row=6, column=col, value=h))
        ws6.row_dimensions[6].height = 30

    # FP K1
    ws6.cell(row=7, column=1, value='K1')
    sc(ws6.cell(row=7, column=1))
    ws6.cell(row=7, column=2, value='=03_Kolom!D5')
    sr(ws6.cell(row=7, column=2))
    ws6.cell(row=7, column=2).number_format = '0.00'
    ws6.cell(row=7, column=3, value='=$B$4')
    sc(ws6.cell(row=7, column=3))
    ws6.cell(row=7, column=4, value=1.0)
    si(ws6.cell(row=7, column=4))
    ws6.cell(row=7, column=4).number_format = '0.00'

    # Luas perlu
    ws6.cell(row=7, column=5, value='=(B7+D7^2*0.4*24+D7^2*0.6*18)/C7')
    sr(ws6.cell(row=7, column=5))
    ws6.cell(row=7, column=5).number_format = '0.00'

    ws6.cell(row=7, column=6, value='=D7^2')
    sc(ws6.cell(row=7, column=6))
    ws6.cell(row=7, column=6).number_format = '0.00'

    ws6.cell(row=7, column=7, value='=B7/F7')
    sr(ws6.cell(row=7, column=7))
    ws6.cell(row=7, column=7).number_format = '0.00'

    ws6.cell(row=7, column=8, value='=IF(G7<=C7,"OK","NOK")')
    st(ws6.cell(row=7, column=8))

    # b0
    ws6.cell(row=7, column=9, value='=4*D7*1000')
    sc(ws6.cell(row=7, column=9))

    # φVn
    ws6.cell(row=7, column=10, value='=01_Input!B28*1/3*SQRT(01_Input!B25*1000000)*I7*350/1000')
    sr(ws6.cell(row=7, column=10))
    ws6.cell(row=7, column=10).number_format = '0.00'

    ws6.cell(row=7, column=11, value='=IF(B7*1000<J7,"OK","NOK")')
    st(ws6.cell(row=7, column=11))

    # FP K2
    ws6.cell(row=8, column=1, value='K2')
    sc(ws6.cell(row=8, column=1))
    ws6.cell(row=8, column=2, value='=03_Kolom!D6')
    sr(ws6.cell(row=8, column=2))
    ws6.cell(row=8, column=2).number_format = '0.00'
    ws6.cell(row=8, column=3, value='=$B$4')
    sc(ws6.cell(row=8, column=3))
    ws6.cell(row=8, column=4, value=1.0)
    si(ws6.cell(row=8, column=4))
    ws6.cell(row=8, column=4).number_format = '0.00'
    ws6.cell(row=8, column=5, value='=(B8+D8^2*0.4*24+D8^2*0.6*18)/C8')
    sr(ws6.cell(row=8, column=5))
    ws6.cell(row=8, column=5).number_format = '0.00'
    ws6.cell(row=8, column=6, value='=D8^2')
    sc(ws6.cell(row=8, column=6))
    ws6.cell(row=8, column=6).number_format = '0.00'
    ws6.cell(row=8, column=7, value='=B8/F8')
    sr(ws6.cell(row=8, column=7))
    ws6.cell(row=8, column=7).number_format = '0.00'
    ws6.cell(row=8, column=8, value='=IF(G8<=C8,"OK","NOK")')
    st(ws6.cell(row=8, column=8))
    ws6.cell(row=8, column=9, value='=4*D8*1000')
    sc(ws6.cell(row=8, column=9))
    ws6.cell(row=8, column=10, value='=01_Input!B28*1/3*SQRT(01_Input!B25*1000000)*I8*350/1000')
    sr(ws6.cell(row=8, column=10))
    ws6.cell(row=8, column=10).number_format = '0.00'
    ws6.cell(row=8, column=11, value='=IF(B8*1000<J8,"OK","NOK")')
    st(ws6.cell(row=8, column=11))

    # FP K3
    ws6.cell(row=9, column=1, value='K3')
    sc(ws6.cell(row=9, column=1))
    ws6.cell(row=9, column=2, value='=03_Kolom!D7')
    sr(ws6.cell(row=9, column=2))
    ws6.cell(row=9, column=2).number_format = '0.00'
    ws6.cell(row=9, column=3, value='=$B$4')
    sc(ws6.cell(row=9, column=3))
    ws6.cell(row=9, column=4, value=1.2)
    si(ws6.cell(row=9, column=4))
    ws6.cell(row=9, column=4).number_format = '0.00'
    ws6.cell(row=9, column=5, value='=(B9+D9^2*0.4*24+D9^2*0.6*18)/C9')
    sr(ws6.cell(row=9, column=5))
    ws6.cell(row=9, column=5).number_format = '0.00'
    ws6.cell(row=9, column=6, value='=D9^2')
    sc(ws6.cell(row=9, column=6))
    ws6.cell(row=9, column=6).number_format = '0.00'
    ws6.cell(row=9, column=7, value='=B9/F9')
    sr(ws6.cell(row=9, column=7))
    ws6.cell(row=9, column=7).number_format = '0.00'
    ws6.cell(row=9, column=8, value='=IF(G9<=C9,"OK","NOK")')
    st(ws6.cell(row=9, column=8))
    ws6.cell(row=9, column=9, value='=4*D9*1000')
    sc(ws6.cell(row=9, column=9))
    ws6.cell(row=9, column=10, value='=01_Input!B28*1/3*SQRT(01_Input!B25*1000000)*I9*350/1000')
    sr(ws6.cell(row=9, column=10))
    ws6.cell(row=9, column=10).number_format = '0.00'
    ws6.cell(row=9, column=11, value='=IF(B9*1000<J9,"OK","NOK")')
    st(ws6.cell(row=9, column=11))

    ws6.column_dimensions['A'].width = 8
    ws6.column_dimensions['B'].width = 10
    ws6.column_dimensions['C'].width = 10
    ws6.column_dimensions['D'].width = 8
    ws6.column_dimensions['E'].width = 12
    ws6.column_dimensions['F'].width = 10
    ws6.column_dimensions['G'].width = 12
    ws6.column_dimensions['H'].width = 10
    ws6.column_dimensions['I'].width = 10
    ws6.column_dimensions['J'].width = 12
    ws6.column_dimensions['K'].width = 12

    # ============================================================
    # SHEET 7: RING BALK - PERHITUNGAN OTOMATIS
    # ============================================================
    ws7 = wb.create_sheet("07_Ring_Balk")

    ws7.merge_cells('A1:H1')
    ws7['A1'] = "PERHITUNGAN RING BALK - OTOMATIS"
    ws7['A1'].font = title_font
    ws7['A1'].alignment = center

    rb_headers = ['Lokasi', 'b (cm)', 'h (cm)', 'Bentang (m)', 'q (kN/m)', 'Mu (kNm)', 'As perlu (mm2)', 'Tul. Utama']
    for col, h in enumerate(rb_headers, 1):
        sh(ws7.cell(row=3, column=col, value=h))

    # Memanjang
    ws7.cell(row=4, column=1, value='Memanjang')
    sc(ws7.cell(row=4, column=1))
    ws7.cell(row=4, column=2, value=15)
    sc(ws7.cell(row=4, column=2))
    ws7.cell(row=4, column=3, value=20)
    sc(ws7.cell(row=4, column=3))
    ws7.cell(row=4, column=4, value='=01_Input!B6')
    sc(ws7.cell(row=4, column=4))
    ws7.cell(row=4, column=4).number_format = '0.00'
    ws7.cell(row=4, column=5, value='=B4*C4*24/1000000+D4*(01_Input!B20+01_Input!B18)/1000')
    sc(ws7.cell(row=4, column=5))
    ws7.cell(row=4, column=5).number_format = '0.00'
    ws7.cell(row=4, column=6, value='=1/10*E4*D4^2')
    sr(ws7.cell(row=4, column=6))
    ws7.cell(row=4, column=6).number_format = '0.00'
    ws7.cell(row=4, column=7, value='=F4*1000000/(01_Input!B28*01_Input!B29*(C4-2.5-0.4)*10)')
    sr(ws7.cell(row=4, column=7))
    ws7.cell(row=4, column=7).number_format = '0.00'
    ws7.cell(row=4, column=8, value='4D10')
    sc(ws7.cell(row=4, column=8))

    # Melintang
    ws7.cell(row=5, column=1, value='Melintang')
    sc(ws7.cell(row=5, column=1))
    ws7.cell(row=5, column=2, value=15)
    sc(ws7.cell(row=5, column=2))
    ws7.cell(row=5, column=3, value=20)
    sc(ws7.cell(row=5, column=3))
    ws7.cell(row=5, column=4, value='=01_Input!B7')
    sc(ws7.cell(row=5, column=4))
    ws7.cell(row=5, column=4).number_format = '0.00'
    ws7.cell(row=5, column=5, value='=B5*C5*24/1000000+D5*(01_Input!B20+01_Input!B18)/1000')
    sc(ws7.cell(row=5, column=5))
    ws7.cell(row=5, column=5).number_format = '0.00'
    ws7.cell(row=5, column=6, value='=1/10*E5*D5^2')
    sr(ws7.cell(row=5, column=6))
    ws7.cell(row=5, column=6).number_format = '0.00'
    ws7.cell(row=5, column=7, value='=F5*1000000/(01_Input!B28*01_Input!B29*(C5-2.5-0.4)*10)')
    sr(ws7.cell(row=5, column=7))
    ws7.cell(row=5, column=7).number_format = '0.00'
    ws7.cell(row=5, column=8, value='4D10')
    sc(ws7.cell(row=5, column=8))

    ws7.column_dimensions['A'].width = 14
    ws7.column_dimensions['B'].width = 8
    ws7.column_dimensions['C'].width = 8
    ws7.column_dimensions['D'].width = 12
    ws7.column_dimensions['E'].width = 12
    ws7.column_dimensions['F'].width = 12
    ws7.column_dimensions['G'].width = 14
    ws7.column_dimensions['H'].width = 12

    # ============================================================
    # SHEET 8: REKAPITULASI & BOM - OTOMATIS
    # ============================================================
    ws8 = wb.create_sheet("08_Rekapitulasi")

    ws8.merge_cells('A1:F1')
    ws8['A1'] = "REKAPITULASI MATERIAL & BOM - OTOMATIS"
    ws8['A1'].font = title_font
    ws8['A1'].alignment = center

    rekap_headers = ['No', 'Komponen', 'Volume (m3)', 'Beton', 'Baja Tul. (kg)', 'Keterangan']
    for col, h in enumerate(rekap_headers, 1):
        sh(ws8.cell(row=3, column=col, value=h))

    # Sloof
    ws8.cell(row=4, column=1, value=1)
    sc(ws8.cell(row=4, column=1))
    ws8.cell(row=4, column=2, value='Sloof')
    sc(ws8.cell(row=4, column=2))
    ws8.cell(row=4, column=3, value='=12*(01_Input!B6+01_Input!B7)*0.15*0.30')
    sc(ws8.cell(row=4, column=3))
    ws8.cell(row=4, column=3).number_format = '0.00'
    ws8.cell(row=4, column=4, value='=C4')
    sc(ws8.cell(row=4, column=4))
    ws8.cell(row=4, column=5, value='=C4*120')
    sr(ws8.cell(row=4, column=5))
    ws8.cell(row=4, column=5).number_format = '0.0'
    ws8.cell(row=4, column=6, value='12 titik')
    sc(ws8.cell(row=4, column=6))

    # Kolom
    ws8.cell(row=5, column=1, value=2)
    sc(ws8.cell(row=5, column=1))
    ws8.cell(row=5, column=2, value='Kolom')
    sc(ws8.cell(row=5, column=2))
    ws8.cell(row=5, column=3, value='=12*01_Input!B8*0.30*0.30')
    sc(ws8.cell(row=5, column=3))
    ws8.cell(row=5, column=3).number_format = '0.00'
    ws8.cell(row=5, column=4, value='=C5')
    sc(ws8.cell(row=5, column=4))
    ws8.cell(row=5, column=5, value='=C5*120')
    sr(ws8.cell(row=5, column=5))
    ws8.cell(row=5, column=5).number_format = '0.0'
    ws8.cell(row=5, column=6, value='12 kolom')
    sc(ws8.cell(row=5, column=6))

    # Balok
    ws8.cell(row=6, column=1, value=3)
    sc(ws8.cell(row=6, column=1))
    ws8.cell(row=6, column=2, value='Balok')
    sc(ws8.cell(row=6, column=2))
    ws8.cell(row=6, column=3, value='=8*(01_Input!B6+01_Input!B7)*0.25*0.40')
    sc(ws8.cell(row=6, column=3))
    ws8.cell(row=6, column=3).number_format = '0.00'
    ws8.cell(row=6, column=4, value='=C6')
    sc(ws8.cell(row=6, column=4))
    ws8.cell(row=6, column=5, value='=C6*120')
    sr(ws8.cell(row=6, column=5))
    ws8.cell(row=6, column=5).number_format = '0.0'
    ws8.cell(row=6, column=6, value='8 balok')
    sc(ws8.cell(row=6, column=6))

    # Plat Lantai
    ws8.cell(row=7, column=1, value=4)
    sc(ws8.cell(row=7, column=1))
    ws8.cell(row=7, column=2, value='Plat Lantai')
    sc(ws8.cell(row=7, column=2))
    ws8.cell(row=7, column=3, value='=01_Input!B6*01_Input!B7*01_Input!B11/100')
    sc(ws8.cell(row=7, column=3))
    ws8.cell(row=7, column=3).number_format = '0.00'
    ws8.cell(row=7, column=4, value='=C7')
    sc(ws8.cell(row=7, column=4))
    ws8.cell(row=7, column=5, value='=C7*120')
    sr(ws8.cell(row=7, column=5))
    ws8.cell(row=7, column=5).number_format = '0.0'
    ws8.cell(row=7, column=6, value='48 m2')
    sc(ws8.cell(row=7, column=6))

    # Ring Balk
    ws8.cell(row=8, column=1, value=5)
    sc(ws8.cell(row=8, column=1))
    ws8.cell(row=8, column=2, value='Ring Balk')
    sc(ws8.cell(row=8, column=2))
    ws8.cell(row=8, column=3, value='=2*(01_Input!B6+01_Input!B7)*0.15*0.20')
    sc(ws8.cell(row=8, column=3))
    ws8.cell(row=8, column=3).number_format = '0.00'
    ws8.cell(row=8, column=4, value='=C8')
    sc(ws8.cell(row=8, column=4))
    ws8.cell(row=8, column=5, value='=C8*120')
    sr(ws8.cell(row=8, column=5))
    ws8.cell(row=8, column=5).number_format = '0.0'
    ws8.cell(row=8, column=6, value='18m')
    sc(ws8.cell(row=8, column=6))

    # Fondasi
    ws8.cell(row=9, column=1, value=6)
    sc(ws8.cell(row=9, column=1))
    ws8.cell(row=9, column=2, value='Fondasi FP')
    sc(ws8.cell(row=9, column=2))
    ws8.cell(row=9, column=3, value='=3*1.0*1.0*0.40')
    sc(ws8.cell(row=9, column=3))
    ws8.cell(row=9, column=3).number_format = '0.00'
    ws8.cell(row=9, column=4, value='=C9')
    sc(ws8.cell(row=9, column=4))
    ws8.cell(row=9, column=5, value='=C9*120')
    sr(ws8.cell(row=9, column=5))
    ws8.cell(row=9, column=5).number_format = '0.0'
    ws8.cell(row=9, column=6, value='3 unit')
    sc(ws8.cell(row=9, column=6))

    # Total
    ws8.merge_cells('A10:B10')
    ws8.cell(row=10, column=1, value='TOTAL')
    sh(ws8.cell(row=10, column=1))
    ws8.cell(row=10, column=3, value='=SUM(C4:C9)')
    sh(ws8.cell(row=10, column=3))
    ws8.cell(row=10, column=3).number_format = '0.00'
    ws8.cell(row=10, column=5, value='=SUM(E4:E9)')
    sh(ws8.cell(row=10, column=5))
    ws8.cell(row=10, column=5).number_format = '0.0'

    ws8.column_dimensions['A'].width = 14
    ws8.column_dimensions['B'].width = 14
    ws8.column_dimensions['C'].width = 12
    ws8.column_dimensions['D'].width = 12
    ws8.column_dimensions['E'].width = 14
    ws8.column_dimensions['F'].width = 12

    # ============================================================
    # SHEET 9: PETUNJUK PENGGUNAAN
    # ============================================================
    ws9 = wb.create_sheet("09_Petunjuk")

    ws9.merge_cells('A1:E1')
    ws9['A1'] = "PETUNJUK PENGGUNAAN SPREADSHEET"
    ws9['A1'].font = Font(bold=True, size=16, color='4472C4')
    ws9['A1'].alignment = center

    ws9.merge_cells('A3:E3')
    ws9['A3'] = "1. UBAH NILAI INPUT DI SHEET '01_Input' (SEL KUNING)"
    ws9['A3'].font = Font(bold=True, size=12)

    ws9.merge_cells('A4:E4')
    ws9['A4'] = "2. HASIL PERHITUNGAN AKAN BERUBAH OTOMATIS DI SETIAP SHEET"

    ws9.merge_cells('A5:E5')
    ws9['A5'] = "3. STATUS 'OK' BERARTI KAPASITAS MEMENUHI SYARAT"

    ws9.merge_cells('A6:E6')
    ws9['A6'] = "4. STATUS 'NOK' BERARTI PERLU DITAMBAH DIMENSI/TULANGAN"

    ws9.merge_cells('A8:E8')
    ws9['A8'] = "REFERENSI SNI:"
    ws9['A8'].font = Font(bold=True)

    sni_ref = [
        'SNI 2847:2019 - Tata cara perencanaan struktur beton',
        'SNI 1726:2019 - Tata cara perencanaan ketahanan gempa',
        'SNI 1727:2020 - Beban desain minimum untuk bangunan gedung',
        'SNI 8460:2017 - Persyaratan perancangan geoteknik',
    ]

    for i, ref in enumerate(sni_ref, 9):
        ws9.cell(row=i, column=1, value=ref).fill = calc_fill
        ws9.cell(row=i, column=1).border = thin_border

    ws9.column_dimensions['A'].width = 50

    # Save
    output_path = r"C:\Users\dinas\OneDrive\Desktop\kode\web-admin\perhitungan_struktur_otomatis.xlsx"
    wb.save(output_path)
    print("Spreadsheet dengan formula Excel berhasil disimpan:", output_path)
    print("\nFITUR:")
    print("- Ubah nilai di sheet '01_Input' (sel kuning)")
    print("- Semua perhitungan otomatis berubah")
    print("- Status OK/NOK untuk setiap komponen")
    print("- Rekapitulasi material di sheet '08_Rekapitulasi'")
    return output_path

if __name__ == "__main__":
    create_calculated_structure()